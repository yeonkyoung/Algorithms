from sklearn.metrics import confusion_matrix, roc_auc_score, classification_report
from tabulate import tabulate
import pandas as pd
import numpy as np
from openpyxl import Workbook

class model_result:
    def __init__(self, model_result_dict):
        self.len_train = model_result_dict['len_train']
        self.len_test  = model_result_dict['len_test']
        self.tr_classes = model_result_dict['tr_classes']
        self.train_num_entity = model_result_dict['train_num_entity']
        self.test_classes = model_result_dict['test_classes']
        self.test_num_entity = model_result_dict['test_num_entity']
        self.cm = model_result_dict['cm']
        self.quality_ind = model_result_dict['quality_ind']
        self.productivity_ind = model_result_dict['productivity_ind']
        self.acc = model_result_dict['acc']
        self.auroc = model_result_dict['auroc']
    
    def make_excel_file(self, file_path):

        write_wb = Workbook()
        write_ws = write_wb.active

        # train set 개수
        write_ws['A2'] = 'Train Set 개수'
        write_ws.merge_cells('A2:A3')
        write_ws['B2'] = '불량'
        write_ws['C2'] = self.train_num_entity[0]
        write_ws['B3'] = '정상'
        write_ws['C3'] = self.train_num_entity[1]
        write_ws['B4'] = 'Train set 합계'
        write_ws['C4'] = self.train_num_entity[0] + self.train_num_entity[1]
        
        # test set 개수
        write_ws['A5'] = 'Test Set 개수'
        write_ws.merge_cells('A5:A6')
        write_ws['B5'] = '불량'
        write_ws['C5'] = self.test_num_entity[0]
        write_ws['B6'] = '정상'
        write_ws['C6'] = self.test_num_entity[1]
        write_ws['B7'] = 'Test set 합계'
        write_ws['C7'] = self.test_num_entity[0] + self.test_num_entity[1]

        #전체 합계
        write_ws['B8'] = '전체 합계'
        write_ws['C8'] = write_ws['C4'].value + write_ws['C7'].value

        # confusion matrix
        write_ws.merge_cells('E10:F11')
        write_ws['E12'] = '실제'
        write_ws.merge_cells('E12:E13')
        write_ws['F12'] = '정상'
        write_ws['F13'] = '불량'
        write_ws['F14'] = '합계'
        write_ws.merge_cells('E14:F14')

        write_ws['G10'] = '예측'
        write_ws.merge_cells('G10:H10')
        write_ws['G11'] = '정상'
        write_ws['H11'] = '불량'
        write_ws['I11'] = '합계'
        write_ws.merge_cells('I10:I11')

        write_ws['G12'] = self.cm.iloc[0][0] # 실제 정상, 예측 정상
        write_ws['H12'] = self.cm.iloc[0][1] # 실제 정상, 예측 불량
        write_ws['G13'] = self.cm.iloc[1][0] # 실제 불량, 예측 정상
        write_ws['H13'] = self.cm.iloc[1][1] # 실제 불량, 예측 불량
        write_ws['I12'] = write_ws['G12'].value + write_ws['H12'].value



        # write_ws['A1'] = self.len_train
        write_wb.save(file_path)


def multi_model_result(y_train, y_test, y_pred):

    ### 데이터 현황
    len_train = len(y_train)
    len_test = len(y_test)
    normal = (y_test == y_pred).sum()
    notnormal = (y_test != y_pred).sum()
    acc = normal / (normal + notnormal) * 100

    ### Confusion Matrix
    cm = pd.DataFrame(confusion_matrix(y_test, y_pred), columns = [1, 2, 3, 4])
    cm.rename(index={0: 1, 1: 2, 2: 3, 3: 4}, inplace = True)

    ### Classification Report
    report = pd.DataFrame(classification_report(y_pred, y_test, target_names = [1, 2, 3, 4], digits = 4, output_dict = True))
    report.rename(columns = {"accuracy" : "정확도"}, inplace = True)
    report.rename(index = {'precision' : "정밀도", "recall" : "재현율"}, inplace = True)
    report = report.transpose().iloc[:4,:3]


    print('### 학습 데이터 총 개수 : {}'.format(len_train))
    print('### 테스트 데이터 총 개수 : {}'.format(len_test)); print()
    print('-- 정상 : {} / {}'.format(normal, len_test))
    print('-- 불량 : {} / {}'.format(notnormal, len_test)); print()
    print('-- 정확도 : {:.2f}'.format(acc))

    print();print()


    print('### Confusion Matrix')
    print(tabulate(cm, headers='keys', tablefmt='grid')) # tablefmt='psql'

    print(); print()


    print('### Classification Report')
    print(tabulate(report, headers='keys', tablefmt='grid')) # tablefmt='psql'

    print(); print()

    return



def make_excel_file(file_path, func):
    def decorater(func):
        def wrapper(y_train, y_test, y_pred):
            len_train, cm, quality_ind, productivity_ind, acc, auroc, train_num_entity, len_test, normal, notnormal = func(y_train, y_test, y_pred) 

            write_wb = Workbook()

            write_ws = write_wb.active
            write_ws['A1'] = len_train
            write_wb.save(file_path)

        return wrapper
    return decorater

# @make_excel_file('test.xlsx')
def binary_model_result(y_train, y_test, y_pred):

    ### 데이터 현황
    len_train = len(y_train)
    len_test = len(y_test)
    # normal = (y_test == y_pred).sum()
    # notnormal = (y_test != y_pred).sum()
    # acc = normal / (normal + notnormal) * 100

    tr_classes, train_num_entity = np.unique(y_train, return_counts=True)
    test_classes, test_num_entity = np.unique(y_test, return_counts=True)
    normal = test_num_entity[0]
    notnormal = test_num_entity[1]
    
    ### Confusion Matrix
    cm = pd.DataFrame(confusion_matrix(y_test, y_pred), columns = ['정상', '불량'])
    cm.rename(index={0: "정상", 1: "불량"}, inplace = True)
    
    ### Classification Report
    report = pd.DataFrame(classification_report(y_pred, y_test, target_names = ['정상', '불량'], digits = 4, output_dict = True))
    report.rename(columns = {"accuracy" : "정확도"}, inplace = True)
    report.rename(index = {'precision' : "정밀도", "recall" : "재현율"}, inplace = True)
    report = report.transpose().iloc[:3,:3]
    
    ### 품질, 생산성 지표
    quality_ind = cm.iloc[1,0] / (cm.iloc[1,0] + cm.iloc[1,1])
    productivity_ind = cm.iloc[0,1] / (cm.iloc[0,1] + cm.iloc[1,1])
    acc = (cm.iloc[0,0]+cm.iloc[1,1])/(cm.iloc[0,0]+cm.iloc[1,0]+cm.iloc[0,1]+cm.iloc[1,1])
    
    ### AUROC
    auroc = roc_auc_score(y_test, y_pred)

    
    # print('### 학습 데이터 총 개수 : {}'.format(len_train))
    # print('-- 학습 정상 : {} / {}'.format(train_num_entity[0], len_train))
    # print('-- 학습 불량 : {} / {}'.format(train_num_entity[1], len_train)); print()
    # print('### 테스트 데이터 총 개수 : {}'.format(len_test)); print()
    # print('-- 테스트 정상 : {} / {}'.format(normal, len_test))
    # print('-- 테스트 불량 : {} / {}'.format(notnormal, len_test)); print()
    # print('-- 정확도 : {:.4f}'.format(acc))

    # print();print()


    # print('### Confusion Matrix')
    # print(tabulate(cm, headers='keys', tablefmt='grid')) # tablefmt='psql'

    # print(); print()


    # print('### Classification Report')
    # print(tabulate(report, headers='keys', tablefmt='grid')) # tablefmt='psql'

    # print(); print()


    # print('### 품질, 생산성 지표'); print()
    # print('-- 미검율 : {}'.format(quality_ind))
    # print('-- 과검율 : {}'.format(productivity_ind))

    # print(); print()


    # print('### AUROC(ROC Curve 면적)'); print()
    # print('-- AUROC : {:.7f}'.format(auroc))
    
    # return len_train, cm, quality_ind, productivity_ind, acc, auroc, train_num_entity, len_test, normal, notnormal
    result_dict = {'len_train': len_train
                  ,'len_test': len_test
                  ,'cm': cm
                  ,'quality_ind': quality_ind
                  ,'productivity_ind': productivity_ind
                  ,'acc': acc
                  ,'auroc': auroc
                  ,'tr_classes': tr_classes
                  ,'test_classes': test_classes
                  ,'train_num_entity': train_num_entity
                  ,'test_num_entity': test_num_entity
                  ,'normal': normal
                  ,'notnormal': notnormal
                  }

    return result_dict